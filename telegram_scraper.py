# Base
import asyncio
import aiohttp
import random

# HTML parsing
from bs4 import BeautifulSoup

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font

# Custom functions
from db_functions import get_user_role


async def fetch_html_content(session, url):
    """
    Fetches HTML content from a given URL.

    Args:
        session: aiohttp ClientSession object.
        url (str): URL to fetch.

    Returns:
        str: HTML content.
    """

    async with session.get(url) as response:
        return await response.text()


async def search_telegram_groups_and_chats(session, keyword, wb, ws, user_id):
    """
    Searches for Telegram groups and chats based on a keyword.

    Args:
        session: aiohttp ClientSession object.
        keyword (str): Keyword to search for.
        wb: openpyxl Workbook object for writing data.
        ws: openpyxl Worksheet object for writing data.
        user_id (int): ID of the user initiating the search.

    Returns:
        None. Writes search results to the Excel file.
    """

    # Retrieving user role for random delay calculation
    role = await get_user_role(user_id)

    # Function for introducing random delays
    async def random_delay():
        if role == "premium":
            await asyncio.sleep(random.uniform(2, 3))
        else:
            await asyncio.sleep(random.uniform(4, 6))

    processed_chat_links = set()
    processed_urls = set()
    group_data = []
    chat_data = []
    page_number = 1

    while True:
        await random_delay()
        tgram_url = f"https://tgsearch.org/search?query={keyword}&page={page_number}"
        async with session.get(
            tgram_url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
            },
        ) as response:
            if response.status != 200:
                break

            html = await response.text()
            soup = BeautifulSoup(html, "html.parser")
            group_containers = soup.find_all("div", class_="channel-card")
            new_groups_found = False

            for container in group_containers:
                ul_element = container.find("ul", class_="channel-card__options")
                li_elements = ul_element.find_all("li")
                name = li_elements[1].text.strip()
                if not name.startswith("@"):
                    continue
                name_replace = name.replace("@", "")
                group_name = container.find("h2", class_="channel-card__title")
                users = li_elements[0].text.strip()
                group_name_text = group_name.a.text.strip()
                url = f"https://t.me/{name_replace}"

                if url in processed_urls:
                    continue

                processed_urls.add(url)
                new_groups_found = True

                group_data.append([group_name_text, url, users, ""])
            if not new_groups_found:
                break

        page_number += 1

    page_number = 1
    processed_chat_links = set()
    while True:
        tgram_url = f"https://tgpulse.com/ru/search?search={keyword}&page={page_number}&lang=&category=&order=online"
        async with session.get(
            tgram_url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
            },
        ) as response:
            if response.status != 200:
                break

            html = await response.text()
            soup = BeautifulSoup(html, "html.parser")
            meta_tags = soup.find_all("meta", itemprop="url")
            group_containers = soup.find_all("div", class_="card-body")

            if not meta_tags:
                break

            if not group_containers:
                break

            tasks = []

            for meta_tag in meta_tags:
                chat_link = meta_tag.get("content")

                if chat_link in processed_chat_links:
                    continue

            for container in group_containers:
                group_link_element = container.find("a")
                if group_link_element:
                    group_name_element = group_link_element.find("h3")
                    if group_name_element:
                        group_name = group_name_element.text.strip()
                participants_container = container.find_all("div", class_="col-6")
                for participant in participants_container:
                    label = participant.find("label")
                    if label and label.text.strip() == "Всего участников":
                        participants = participant.find("h4").text.strip()

                        chat_data.append([group_name, chat_link, participants])

            if not tasks:
                break

            await asyncio.sleep(random.uniform(1, 3))
        page_number += 1

    combined_data = []
    max_len = max(len(group_data), len(chat_data))
    for i in range(max_len):
        group_row = group_data[i] if i < len(group_data) else [""] * len(group_data[0])
        chat_row = chat_data[i] if i < len(chat_data) else [""]
        combined_data.append(group_row + chat_row)

    for row in combined_data:
        ws.append(row)


async def fetch_telegram_data(user_id, keywords):
    """
    Searches for Telegram groups and chats based on given keywords.

    Args:
        user_id (int): ID of the user initiating the search.
        keywords (str): Keywords separated by commas.

    Returns:
        None. Writes search results to an Excel file.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Telegram Data"
    ws.append(["Каналы:", "", "", "", "Чаты:", "", "", ""])
    ws.append(["Название", "URL", "Участники ", "", "Название", "URL", "Участники "])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 27
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 10

    header_style = Font(name="Arial", size=12, bold=True)
    title_style = Font(name="Arial", size=10, bold=True)
    ws.cell(row=1, column=1).font = header_style
    ws.cell(row=1, column=5).font = header_style
    ws.cell(row=2, column=1).font = title_style
    ws.cell(row=2, column=2).font = title_style
    ws.cell(row=2, column=3).font = title_style
    ws.cell(row=2, column=5).font = title_style
    ws.cell(row=2, column=6).font = title_style
    ws.cell(row=2, column=7).font = title_style

    async with aiohttp.ClientSession() as session:
        for keyword in keywords.split(","):
            await search_telegram_groups_and_chats(
                session, keyword.strip(), wb, ws, user_id
            )
    wb.save("group_chat_data.xlsx")
