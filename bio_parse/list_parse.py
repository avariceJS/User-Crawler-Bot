# Async
import asyncio
import aiohttp

# Data manipulation
import openpyxl
from datetime import datetime, timezone
import random

# HTML parsing
from bs4 import BeautifulSoup

# Excel
from openpyxl.styles import Font

# Telethon
from telethon.tl.functions.channels import GetFullChannelRequest
from telethon.tl.types import UserStatusRecently, UserStatusOnline, UserStatusOffline
from telethon_client import get_telethon_client

# Custom functions
from db_functions import get_user_role


async def bio_parse_from_group(url, user_id) -> None:
    """
    Parse bio data of Telegram users from a group.

    Args:
        url (str): URL of the Telegram group.
        user_id (int): ID of the user initiating the parsing.

    Returns:
        None. Writes parsed data to an Excel file.
    """

    group_url = url

    try:

        # Establishing a connection with the Telegram client
        client = await get_telethon_client()
        target_group = await client(GetFullChannelRequest(channel=group_url))
        all_participants = await client.get_participants(target_group.full_chat.id)

        # Determine the parsing limit based on user role
        role = await get_user_role(user_id)
        if role == "Базовая":
            num_to_parse = min(130, len(all_participants))
        else:
            num_to_parse = min(200, len(all_participants))

        # Randomly select users to parse from the group
        users_to_parse = random.sample(all_participants, num_to_parse)

        # Fetching user descriptions concurrently
        async with aiohttp.ClientSession() as session:
            tasks = []
            for user in users_to_parse:
                username = user.username if user.username else "__"
                url = f"https://t.me/{username}"
                tasks.append(fetch_description(session, url))
            descriptions = await asyncio.gather(*tasks)

        # Creating a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "Имя",
                "Фамилия",
                "Никнейм",
                "Номер телефона",
                "Описание профиля",
                "Дата последней активности",
                " ",
            ]
        )

        # Setting column widths and title font
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 35
        ws.column_dimensions["F"].width = 30
        title_style = Font(name="Arial", size=10, bold=True)
        for col in range(1, 7):
            ws.cell(row=1, column=col).font = title_style

        # Writing user data to the Excel worksheet
        for user, description in zip(users_to_parse, descriptions):
            first_name = user.first_name if user.first_name else "__"
            last_name = user.last_name if user.last_name else "__"
            username = user.username if user.username else "__"
            phone_number = user.phone if user.phone else "__"

            # Determining user's last activity status
            if isinstance(user.status, UserStatusRecently):
                last_seen = "Недавно онлайн"
            else:
                if isinstance(user.status, UserStatusOnline):
                    last_seen = "Онлайн"
                elif isinstance(user.status, UserStatusOffline):
                    last_seen = user.status.was_online
                    current_time = datetime.now(timezone.utc)
                    time_difference = current_time - last_seen
                    minutes_difference = int(time_difference.total_seconds() / 60)
                    hours_difference = minutes_difference // 60
                    remaining_minutes = minutes_difference % 60
                    if hours_difference > 0:
                        last_seen = f"был(а): {hours_difference} часов {remaining_minutes} минут назад"
                    else:
                        last_seen = f"был(а): {remaining_minutes} минут назад"
                else:
                    last_seen = ""

            ws.append(
                [first_name, last_name, username, phone_number, description, last_seen]
            )

        # Formatting cells to allow text wrapping
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Saving the Excel file
        wb.save("user_info.xlsx")

        return "user_info.xlsx"

    except Exception as e:
        return f"Произошла ошибка: {e}"


async def fetch_description(session, url):
    """
    Fetch user description from Telegram web page.

    Args:
        session: aiohttp ClientSession object.
        url (str): URL of the user's Telegram profile.

    Returns:
        str: User's profile description.
    """

    async with session.get(url) as response:
        if response.status == 200:
            response_text = await response.text()
            soup = BeautifulSoup(response_text, "html.parser")

            # Extracting user description from the HTML page
            description_elem = soup.find("div", class_="tgme_page_description")
            description = (
                description_elem.get_text(strip=True)
                if description_elem
                else "Нет описания профиля"
            )
            if not description:
                username_elem = soup.find("div", class_="tgme_page_extra")
                if username_elem and "telegram.me/" in username_elem.text:
                    description = (
                        username_elem.text.strip()
                        .replace("If you haveTelegram, you can contact ", "")
                        .replace(" right away.", "")
                    )
            return description
        else:
            return "Нет описания профиля"
