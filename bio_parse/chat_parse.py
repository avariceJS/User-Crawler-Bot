# Async
import asyncio
import aiohttp

# HTML parsing
from bs4 import BeautifulSoup

# Custom functions
from telethon_client import get_telethon_client
from db_functions import get_user_role

# Excel
import openpyxl
from openpyxl.styles import Font

# Date
from datetime import datetime, timezone

# Telethon
from telethon.tl.functions.messages import GetHistoryRequest
from telethon.tl.types import (
    PeerUser,
    UserStatusRecently,
    UserStatusOnline,
    UserStatusOffline,
)


async def members_id(group_link, limit, user_id):
    """
    Retrieve user IDs from a Telegram group.

    Args:
        group_link (str): Link to the Telegram group.
        limit (int): Maximum number of messages to fetch.

    Returns:
        set: Set of unique user IDs.
    """

    role = await get_user_role(user_id)
    async def delay():
        if role == "premium":
            await asyncio.sleep(23)
        else:
            await asyncio.sleep(33)
    client = await get_telethon_client()
    entity = await client.get_entity(group_link)
    user_ids = set()
    message_count = 0
    has_more_messages = True
    last_message_id = 0

    while has_more_messages and message_count < limit:
        messages = await client(
            GetHistoryRequest(
                peer=entity,
                limit=100,
                offset_date=None,
                offset_id=last_message_id,
                max_id=0,
                min_id=0,
                add_offset=0,
                hash=0,
            )
        )

        for message in messages.messages:
            if isinstance(message.from_id, PeerUser):
                user_ids.add(message.from_id.user_id)
                message_count += 1
                if message_count >= limit:
                    break
                if message_count % 300 == 0:
                    await delay()

        if messages.messages:
            last_message_id = messages.messages[-1].id
        has_more_messages = len(messages.messages) > 0

    await client.disconnect()

    return user_ids


async def bio_parse(user_ids):
    """
    Parse user information from their IDs.

    Args:
        user_ids (set): Set of user IDs.

    Returns:
        None. Writes parsed data to an Excel file.
    """

    client = await get_telethon_client()
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_description(session, client, user_id) for user_id in user_ids]
        descriptions = await asyncio.gather(*tasks)

        # Creating a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "Имя",
                "Фамилия",
                "Никнейм",
                "Номер телефона",
                "Описание профиля",
                "Дата последней активности",
            ]
        )

        # Setting column widths and title font
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 35
        ws.column_dimensions["F"].width = 30
        title_style = Font(name="Arial", size=10, bold=True)
        for col in range(1, 7):
            ws.cell(row=1, column=col).font = title_style

        # Writing user data to the Excel worksheet
        for user_data in descriptions:
            ws.append(user_data)

        # Formatting cells to allow text wrapping
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Saving the Excel file
        wb.save("user_info.xlsx")

    await client.disconnect()


async def fetch_description(session, client, user_id):
    """
    Fetch user description from Telegram web page.

    Args:
        session: aiohttp ClientSession object.
        client: Telethon client.
        user_id (int): ID of the user.

    Returns:
        list: User's profile information.
    """

    try:
        user = await client.get_entity(user_id)
        username = user.username if user.username else "None"
        url = f"https://t.me/{username}" if username != "None" else None

        description = "Нет описания профиля"
        if url:
            async with session.get(url) as response:
                if response.status == 200:
                    response_text = await response.text()
                    soup = BeautifulSoup(response_text, "html.parser")
                    description_elem = soup.find("div", class_="tgme_page_description")
                    description = (
                        description_elem.get_text(strip=True)
                        if description_elem
                        else "Нет описания профиля"
                    )

        first_name = user.first_name if user.first_name else "__"
        last_name = user.last_name if user.last_name else "__"
        phone_number = user.phone if user.phone else "__"

        if isinstance(user.status, UserStatusRecently):
            last_seen = "Недавно онлайн"
        elif isinstance(user.status, UserStatusOnline):
            last_seen = "Онлайн"
        elif isinstance(user.status, UserStatusOffline):
            last_seen = user.status.was_online
            current_time = datetime.now(timezone.utc)
            time_difference = current_time - last_seen
            minutes_difference = int(time_difference.total_seconds() / 60)
            hours_difference = minutes_difference // 60
            remaining_minutes = minutes_difference % 60
            if hours_difference > 0:
                last_seen = (
                    f"был(а): {hours_difference} часов {remaining_minutes} минут назад"
                )
            else:
                last_seen = f"был(а): {remaining_minutes} минут назад"
        else:
            last_seen = ""

        return [first_name, last_name, username, phone_number, description, last_seen]

    except Exception as e:
        return "Ошибка при парсинге"


async def bio_parse_group_chat(group_link, limit, user_id):
    """
    Parse user information from a Telegram group.

    Args:
        group_link (str): Link to the Telegram group.
        limit (int): Maximum number of messages to fetch.

    Returns:
        None. Writes parsed data to an Excel file.
    """
    user_ids = await members_id(group_link, limit, user_id)
    await bio_parse(user_ids)
