# Base
import re

# Excel
import openpyxl
from openpyxl.styles import Font


def extract_minutes(last_seen_value):
    """
    Extract the minutes from the last seen value.

    Args:
        last_seen_value (str): The string indicating the last seen status.

    Returns:
        dict: A dictionary with line numbers as keys and minutes as values.
    """

    minutes_dict = {}
    lines = last_seen_value.split("\n")
    for line in lines:
        if "Онлайн" in line:
            minutes_dict[len(minutes_dict) + 1] = -1
        elif "бы(а):" in line:
            parts = line.split(":")
            if len(parts) == 2:
                time_str = parts[1].strip()
                if "минут" in time_str:
                    minutes = int(time_str.split()[0])
                    minutes_dict[len(minutes_dict) + 1] = minutes
    return minutes_dict


def time_str_to_minutes(time_str):
    """
    Convert time string to minutes.

    Args:
        time_str (str): Time string (e.g., "2 часа 30 минут").

    Returns:
        int: Total number of minutes.
    """

    minutes = 0
    hours_match = re.search(r"(\d+)\s*час", time_str)
    if hours_match:
        hours = int(hours_match.group(1))
        minutes += hours * 60

    minutes_match = re.search(r"(\d+)\s*минут", time_str)
    if minutes_match:
        minutes += int(minutes_match.group(1))

    if not hours_match and not minutes_match:
        minutes_only_match = re.search(r"(\d+)\s*мин", time_str)
        if minutes_only_match:
            minutes += int(minutes_only_match.group(1))
    return minutes


def filter_users(file_path, keep_recently_active):
    """
    Filter users based on their last activity time.

    Args:
        file_path (str): Path to the input Excel file.
        keep_recently_active (bool): Whether to keep users recently active.

    Returns:
        str or None: Path to the filtered Excel file or None if an error occurred.
    """

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        minutes_dict = {}

        last_seen_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Дата последней активности":
                last_seen_col = col
                break

        if last_seen_col is None:
            raise ValueError("Колонка 'Дата последней активности' не найдена в файле")

        for row in range(2, ws.max_row + 1):
            last_seen_value = ws.cell(row=row, column=last_seen_col).value
            if last_seen_value is not None:
                if "Недавно" in last_seen_value:
                    if keep_recently_active:
                        minutes_dict[row] = 99999
                else:
                    if "Онлайн" in last_seen_value:
                        minutes_dict[row] = -1
                    elif "был(а):" in last_seen_value:
                        time_str = last_seen_value.split(": ")[1]
                        minutes = time_str_to_minutes(time_str)
                        minutes_dict[row] = minutes
                    else:
                        lines = last_seen_value.split("\n")
                        for line in lines:
                            if "был(а):" in line:
                                parts = line.split(":")
                                if len(parts) == 2:
                                    time_str = parts[1].strip()
                                    if "минут" in time_str:
                                        minutes = int(time_str.split()[0])
                                        minutes_dict[row] = minutes
                                break

        sorted_d = dict(sorted(minutes_dict.items(), key=lambda item: item[1]))
        sorted_order = sorted(sorted_d.items(), key=lambda x: x[1])
        order_list = [item[0] - 0 for item in sorted_order]
        filter_file_path = file_path.replace(".xlsx", "_filter.xlsx")
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        for col in range(1, ws.max_column + 1):
            new_ws.cell(row=1, column=col).value = ws.cell(row=1, column=col).value

        title_style = Font(name="Arial", size=10, bold=True)
        column_widths = {1: 20, 2: 16, 3: 18, 4: 16, 5: 35, 6: 30}

        for col, width in column_widths.items():
            column_letter = openpyxl.utils.get_column_letter(col)
            new_ws.column_dimensions[column_letter].width = width
            new_ws.cell(row=1, column=col).font = title_style

        for idx, row_idx in enumerate(order_list, start=2):
            for col in range(1, ws.max_column + 1):
                new_ws.cell(row=idx, column=col).value = ws.cell(
                    row=row_idx, column=col
                ).value

        new_wb.save(filter_file_path)
        return filter_file_path

    except Exception as e:
        return f"Произошла ошибка: {e}"
